import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook

# 定义学生类
class Student:
    def __init__(self, student_id, name):
        self.student_id = student_id
        self.name = name
        self.PPT = 0
        self.WORD = 0
        self.EXCEL = 0
        self.basic_operation = 0
        self.errors = 162

def modify_or_add(studentParam):
    found = False
    for student in collection_A:
        # 找到学生,则需要修改值.
        if (student.student_id == studentParam[0]) and (student.name ==studentParam[1]):
            found = True
            if commonValue == "出错数":
                student.errors = studentParam[2]
            elif commonValue == "基础操作":
                student.basic_operation = studentParam[2]
            elif commonValue == "WORD":
                student.WORD =  studentParam[2]
            elif commonValue == "EXCEL":
                student.EXCEL = studentParam[2]
            elif commonValue == "PPT":
               student.PPT = studentParam[2]
            else:
                print("匹配出错!")
                exit()
            # print(f"学生学号为 {studentParam[0]} 的信息已修改。")

    if not found:
        # 没有找到学生,创建对象并添加到集合A中.
        student = Student(studentParam[0], studentParam[1])        
        if commonValue == "出错数":
            student.errors = studentParam[2]
        elif commonValue == "基础操作":
            student.basic_operation = studentParam[2]
        elif commonValue == "WORD":
            student.WORD =  studentParam[2]
        elif commonValue == "EXCEL":
            student.EXCEL = studentParam[2]
        elif commonValue == "PPT":
            student.PPT = studentParam[2]
        else:
            print("匹配出错!")
            exit()
        collection_A.append(student)
        # print(f"学生学号为 {studentParam[0]} 的新信息已添加。")
        

# 定义目标文件夹路径
folder_path = "C:\\Users\\Administrator\\Desktop\\TestMerge\\数字媒体"

# 获取目标文件夹中的所有文件
files = os.listdir(folder_path)

# 过滤出Excel文件
excel_files = [f for f in files if f.endswith('.xlsx')]

# 集合A
collection_A = []

# 遍历Excel文件并逐行打印内容
for file_name in excel_files:
    file_path = os.path.join(folder_path, file_name)
    print(f"读取文件：{file_path}")
    
    # 加载工作簿
    workbook = load_workbook(filename=file_path, data_only=True)
    
    # 获取第一个工作表（可根据需要更改索引或名称）
    sheet = workbook.worksheets[0]
    
    # 逐行打印内容
    commonValue = ""  
    for row in sheet.iter_rows(values_only=True):
        if commonValue == "":
            commonValue = row[2]
            continue
        # print(row[0],row[1],row[2])
        # print(commonValue)
        modify_or_add(row)
    print("\n")

def printMerge():    
    for student in collection_A:
        print(student.student_id,student.name,student.errors,student.basic_operation,student.WORD,student.EXCEL,student.PPT)
        print

data = []
for student in collection_A:
    data.append([student.student_id, student.name, student.errors, student.basic_operation, student.WORD, student.EXCEL, student.PPT])

columns = ["学号", "姓名", "出错数", "基本操作", "WORD", "EXCEL", "PPT"]
df = pd.DataFrame(data, columns=columns)

# 按照学号进行排序
df.sort_values(by='学号', inplace=True)

# 将DataFrame写入Excel文件
output_file = "Merge.xlsx"
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    writer.book = Workbook()
    df.to_excel(writer, sheet_name='Student信息表', index=False)

print("写入完成!")